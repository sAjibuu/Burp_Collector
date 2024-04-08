# Importing necessary modules
import re
import xml.etree.ElementTree as ET
import base64
import json
import os
import multiprocessing
import time
import argparse
import requests
import traceback
import subprocess
import sys
import urllib.parse
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from urllib.parse import urlparse, parse_qs, urlencode
import math
import warnings
import spacy

# Coded by Sagiv Michael

# Defining some colors for output formatting
GREEN = "\033[32m"
RESET = "\033[0m"
RED = "\033[31m"
BLUE = "\033[1;34m"
ORANGE = "\033[1;33m"
MAGENTA = "\033[1;35m"
YELLOW = "\033[33m"


def create_worksheet_count():
    # Creating a new Workbook object
    wb_count = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet_count = wb_count.active

    sheet_count.title = "Exported from Postman"

    sheet_count['A1'] = 'HTTP Methods'
    sheet_count['B1'] = 'Total Endpoints'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet_count['A1'].font = header_font
    sheet_count['B1'].font = header_font

    return wb_count, sheet_count


def create_worksheet_postmantoexcel():
    # Creating a new Workbook object
    wb = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active

    sheet.title = "Exported from Postman"

    sheet['A1'] = 'URL'
    sheet['B1'] = 'Endpoint'
    sheet['C1'] = 'Method'
    sheet['D1'] = 'Description'
    sheet['E1'] = 'Tested?'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font
    sheet['D1'].font = header_font
    sheet['E1'].font = header_font

    return wb, sheet


def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.3
        sheet.column_dimensions[column].width = adjusted_width


HARDCODED_EXTENSIONS = [
    ".jpg", ".jpeg", ".png", ".gif", ".pdf", ".svg", ".json",
    ".css", ".js", ".webp", ".woff", ".woff2", ".eot", ".ttf", ".otf", ".mp4", ".txt"
]


def has_extension(url, extensions):
    """
    Check if the URL has a file extension matching any of the provided extensions.

    Args:
        url (str): The URL to check.
        extensions (list): List of file extensions to match against.

    Returns:
        bool: True if the URL has a matching extension, False otherwise.
    """
    parsed_url = urlparse(url)
    path = parsed_url.path
    extension = os.path.splitext(path)[1].lower()

    return extension in extensions


def clean_url(url):
    """
    Clean the URL by removing redundant port information for HTTP and HTTPS URLs.

    Args:
        url (str): The URL to clean.

    Returns:
        str: Cleaned URL.
    """
    parsed_url = urlparse(url)

    if (parsed_url.port == 80 and parsed_url.scheme == "http") or (
            parsed_url.port == 443 and parsed_url.scheme == "https"):
        parsed_url = parsed_url._replace(netloc=parsed_url.netloc.rsplit(":", 1)[0])

    return parsed_url.geturl()


def clean_urls(urls, extensions):
    """
    Clean a list of URLs by removing unnecessary parameters and query strings.

    Args:
        urls (list): List of URLs to clean.
        extensions (list): List of file extensions to check against.

    Returns:
        list: List of cleaned URLs.
    """

    cleaned_urls = set()

    for url in urls:
        cleaned_url = clean_url(url)
        if not has_extension(cleaned_url, extensions):
            parsed_url = urlparse(cleaned_url)
            query_params = parse_qs(parsed_url.query)
            cleaned_params = {key: "FUZZ" for key in query_params}
            cleaned_query = urlencode(cleaned_params, doseq=True)
            cleaned_url = parsed_url._replace(query=cleaned_query).geturl()
            cleaned_urls.add(cleaned_url)

    return list(cleaned_urls)


def fetch_and_clean_urls(domain, extensions, urls, host):
    tmp_cleaned_urls = clean_urls(urls, extensions)
    cleaned_urls = []

    for clean_url in tmp_cleaned_urls:
        if "FUZZ" in clean_url:
            cleaned_urls.append(clean_url)

    print(f"\n{YELLOW}[INFO]{RESET} Cleaning URLs for: {BLUE}{domain}{RESET}")
    print(f"\n{YELLOW}[INFO]{RESET} Found {GREEN + str(len(cleaned_urls)) + RESET} URLs after cleaning")

    results_dir = host
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    parts = host.split('.')
    # Join the parts without the last part (the TLD)
    domain = '.'.join(parts[:-1])
    domain = domain + "_queries.txt"

    result_file = os.path.join(results_dir, domain)
    with open(result_file, "w") as f:
        for url in cleaned_urls:
            f.write(url + "\n")

    print(f"{YELLOW}[INFO]{RESET} Saved cleaned URLs to {BLUE + result_file + RESET}")
    print(f"{YELLOW}[INFO]{RESET} You can now run fuzzing tests (SQLi, XSS, CRLF, etc') with Nuclei.")
    print(
        f"{YELLOW}[INFO]{RESET} Example: nuclei -t fuzzing-templates -list {domain} -proxy http://{{IPAddress}}:{{PORT}} -H 'Cookie: {{cookies}}' {RESET}")


def paramspider(file):
    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()
    urls = []

    for i in root:
        url = i.find('url').text
        urls.append(url)
        domain = i.find('host').text
        break

    for i in root:
        url = i.find('url').text
        urls.append(url)

    extensions = HARDCODED_EXTENSIONS

    fetch_and_clean_urls(domain, extensions, urls, domain)


def extract_postman_data(json_data):
    for item in json_data.get('item', []):
        if 'request' in item:
            method = item['request']['method'].upper()

            # Check if the 'url' is a string or a dictionary
            if isinstance(item['request']['url'], dict):
                url = item['request']['url'].get('raw', '')
            else:
                url = item['request']['url']

            # Remove {{baseUrl}} from the URL
            url = url.replace("{{baseUrl}}", "").strip()

            # Capture the description
            description = item['request'].get('description', '')

            data_postman.append([base_url, url, method, description])

        if 'item' in item:
            extract_postman_data(item)  # recursive call for folders within collection

def extract_base_url(json_data):
    for variable in json_data.get('variable', []):
        if variable.get('key') == 'baseUrl' or variable.get('key') == 'url':
            return variable.get('value', '')
    return ""

def count_endpoints(json_data):
    # OpenAPI Logic
    if 'info' in json_data and ('openapi' in json_data or '2.0' in str(json_data.get('swagger', ''))):
        paths = json_data.get("paths", {})
        for path, path_data in paths.items():
            for method in path_data.keys():
                method = method.upper()
                method_counter[method] = method_counter.get(method, 0) + 1
    # Postman Logic
    elif 'info' in json_data and json_data['info'].get('schema', '') in [
        "https://schema.getpostman.com/json/collection/v2.0.0/collection.json",
        "https://schema.getpostman.com/json/collection/v2.0.1/collection.json",
        "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
    ]:
        extract_postman_methods(json_data)


def extract_postman_methods(json_data):
    for item in json_data.get('item', []):
        if 'request' in item:
            method = item['request']['method'].upper()
            method_counter[method] = method_counter.get(method, 0) + 1
        if 'item' in item:
            extract_postman_methods(item)  # recursive call for folders within collection


def extract_endpoints(json_data):
    global base_url
    # OpenAPI Logic
    if 'info' in json_data and ('openapi' in json_data or '2.0' in str(json_data.get('swagger', ''))):
        url = ""
        if "'servers':" in str(json_data):
                    url = json_data['servers'][0]['url']
        for path, methods in json_data.get("paths", {}).items():
            
            for method, details in methods.items():
                description = details.get("description", "")
                data_postman.append([url, path, method.upper(), description])

    # Postman Logic
    elif 'info' in json_data and json_data['info'].get('schema',
                                                       '') == "https://schema.getpostman.com/json/collection/v2.0.0/collection.json" \
            or json_data['info'].get('schema',
                                     '') == "https://schema.getpostman.com/json/collection/v2.0.1/collection.json" \
            or json_data['info'].get('schema',
                                     '') == "https://schema.getpostman.com/json/collection/v2.1.0/collection.json":
        base_url = extract_base_url(json_data)
        extract_postman_data(json_data)


def postmanDirectory(json_directory, count=None):
    # Sorting and removing duplicates
    counter = 0

    global data_postman
    global data_methods
    global method_counter
    global unique_endpoint

    wb_all, sheet_all = create_worksheet_postmantoexcel()

    for filename in os.listdir(json_directory):

        if filename.endswith('.json'):

            file_path = os.path.join(json_directory, filename)

            # Open and load the JSON file
            with open(file_path, encoding='utf-8') as file:
                postman_data = json.load(file)
                counter += 1

                data_postman = []
                data_methods = []
                method_counter = {}
                unique_endpoint = []
                wb, sheet = create_worksheet_postmantoexcel()
                wb_method, sheet_method = create_worksheet_count()

                if count is not None:
                    count_endpoints(postman_data)
                else:
                    extract_endpoints(postman_data)

                if count is None:
                    # Sorting and removing duplicates
                    final_data = sorted(list(set([tuple(row) for row in data_postman])))

                    for row in final_data:
                        if None in row:
                            continue

                        sheet.append(row)
                        sheet_all.append(row)

                        # Setting font of the cell to Calibri 14
                        row = sheet.max_row
                        all_row = sheet_all.max_row

                        sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
                        sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
                        sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
                        sheet.cell(row=row, column=4).font = Font(name='Calibri', size=14)
                        sheet.cell(row=row, column=5).font = Font(name='Calibri', size=14)

                        sheet_all.cell(row=all_row, column=1).font = Font(name='Calibri', size=14)
                        sheet_all.cell(row=all_row, column=2).font = Font(name='Calibri', size=14)
                        sheet_all.cell(row=all_row, column=3).font = Font(name='Calibri', size=14)
                        sheet_all.cell(row=all_row, column=4).font = Font(name='Calibri', size=14)
                        sheet_all.cell(row=all_row, column=5).font = Font(name='Calibri', size=14)

                    adjust_column_widths(sheet)
                    filename = filename.replace(".json", "")
                    wb.save(f'{filename}_endpoints.xlsx')
                    print(
                        f"{GREEN}[SUCCESS]{RESET} {ORANGE}{filename}_endpoints.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")


                else:
                    total = 0
                    methods = ""
                    for method, count in method_counter.items():
                        total += count
                        methods += f"{method}: {count}\n"
                        print(f"{YELLOW}[INFO]{RESET} {method}: {GREEN}{count}{RESET}")
                    data_methods.append([methods, total])
                    print(f"{YELLOW}[INFO]{RESET} Total Endpoints: {GREEN}{total}{RESET}")
                    answer = input(
                        f"{YELLOW}[INFO]{RESET} Do you wish to export the results to Excel sheet? {GREEN}Y{RESET}/{RED}N{RESET}: ")

                    if answer.upper() == "Y":

                        # Sorting and removing duplicates
                        for row in data_methods:
                            sheet_method.append(row)
                            # Setting font of the cell to Calibri 14
                            row = sheet_method.max_row
                            sheet_method.cell(row=row, column=1).font = Font(name='Calibri', size=14)
                            sheet_method.cell(row=row, column=2).font = Font(name='Calibri', size=14)
                            sheet_method.cell(row=row, column=2).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
                            sheet_method.cell(row=row, column=3).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')

                        adjust_column_widths(sheet_method)
                        wb_method.save(f'Total_HTTP_Methods_Num-{counter}.xlsx')
                        print(
                            f"{GREEN}[SUCCESS]{RESET} {ORANGE}Total_HTTP_Methods_Num-{counter}.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")

    if count is None:
        adjust_column_widths(sheet_all)
        wb_all.save(f'Postman_All_API_Endpoints.xlsx')
        print(
            f"{GREEN}[SUCCESS]{RESET} {ORANGE}Postman_API_Endpoints.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")


def postmanFile(file, count=None):
    global data_postman
    global data_methods
    global method_counter
    global unique_endpoint

    data_postman = []
    data_methods = []
    method_counter = {}
    unique_endpoint = []

    # Open and load the JSON file
    with open(file, 'r', encoding='utf-8') as file:
        postman_data = json.load(file)

        if count is not None:
            count_endpoints(postman_data)
            wb_method, sheet_method = create_worksheet_count()

        else:
            extract_endpoints(postman_data)
            # Sorting and removing duplicates
            wb, sheet = create_worksheet_postmantoexcel()

    if count is None:

        # Sorting and removing duplicates
        final_data = sorted(list(set([tuple(row) for row in data_postman])))

        for row in final_data:
            if None in row:
                continue
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=4).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=5).font = Font(name='Calibri', size=14)

        adjust_column_widths(sheet)
        wb.save(f'API_Endpoints.xlsx')
        print(
            f"{GREEN}[SUCCESS]{RESET} {ORANGE}API_Endpoints.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")

    else:
        total = 0
        methods = ""
        for method, count in method_counter.items():
            total += count
            methods += f"{method}: {count}\n"
            print(f"{YELLOW}[INFO] {RESET}{method}: {GREEN}{count}{RESET}")
        data_methods.append([methods, total])
        print(f"{YELLOW}[INFO]{RESET} Total Endpoints: {GREEN}{total}{RESET}")
        answer = input(
            f"{YELLOW}[INFO]{RESET} Do you wish to export the results to Excel sheet? {GREEN}Y{RESET}/{RED}N{RESET}: ")

        if answer.upper() == "Y":

            # Sorting and removing duplicates
            for row in data_methods:
                sheet_method.append(row)
                # Setting font of the cell to Calibri 14
                row = sheet_method.max_row
                sheet_method.cell(row=row, column=1).font = Font(name='Calibri', size=14)
                sheet_method.cell(row=row, column=2).font = Font(name='Calibri', size=14)
                sheet_method.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

            adjust_column_widths(sheet_method)
            wb_method.save(f'Total_HTTP_Methods.xlsx')
            print(
                f"{GREEN}[SUCCESS]{RESET} {ORANGE}Total_HTTP_Methods.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")


def cleaning(host, lines):
    try:
        # Installation
        # pip install spacy
        # python -m spacy download en_core_web_sm
        nlp = spacy.load("en_core_web_sm")

        regexes = [
            r".{100,}",  # Ignore lines with more than 100 characters (overly specific)
            r"[0-9]{4,}",  # Ignore lines with 4 or more consecutive digits (likely an id)
            r"[0-9]{3,}$",  # Ignore lines where the last 3 or more characters are digits (likely an id)
            r"[a-z0-9]{32}",  # Likely MD5 hash or similar
            r"\b[A-Z]{2,}\w*\b",  # Matches uppercase strings with two or more characters
            r"\b[A-Z0-9]{5,}\b",  # Matches strings with five or more uppercase letters or digits
            r"\b[A-Z]{2,}\w*\b"  # Matches a word beginning with 2 or more Uppercase letters 
            r"\b[BCD]{3,}\b",  # Matches strings with three or more consecutive B, C, or D characters
            r"[0-9]+[A-Z0-9]{5,}",  # Number followed by 5 or more numbers and uppercase letters (almost all noise)
            r"\/.*\/.*\/.*\/.*\/.*\/.*\/",  # Ignore lines more than 6 directories deep (overly specific)
            r"\w{8}-\w{4}-\w{4}-\w{4}-\w{12}",  # Ignore UUIDs
            r"[0-9]+[a-zA-Z]+[0-9]+[a-zA-Z]+[0-9]+",
            # Ignore multiple numbers and letters mixed together (likely noise)
            r"\.(png|jpg|jpeg|gif|svg|bmp|ttf|avif|wav|mp4|aac|ajax|css|all)$",  # Ignore low-value file types
            r"^$",  # Ignores blank lines
            r"[^a-zA-Z0-9\s_.-]+",
            # Remove non-alphanumeric characters except underscore, dash, and dot at the beginning of a line
        ]

        print(f'\n{YELLOW}[INFO]{RESET} Cleaning Wordlist, please wait this make take a while.')
        print(f'\n{YELLOW}[INFO]{RESET} Crazy A** calculations is happening right now man, chill...')

        original_size = len(lines)

        # Apply regexes to remove lines
        for regex in regexes:
            pattern = re.compile(regex)
            lines = [line for line in lines if not pattern.search(line)]

        # Remove lines starting with digits
        lines = [line for line in lines if not re.search(r"^[0-9]", line)]

        # Remove lines that contain only a single character
        lines = [line for line in lines if len(line.strip()) > 1]

        # Sort and remove duplicates
        lines = sorted(set(lines))

        second_lines = [
            line.replace(".js.map", "").replace(".js", "").replace(".map", "").replace(".min.js", "").replace(
                ".min.map", "").replace(".min", "").replace(".html", "")
            for line in lines
            if ('-' in line or '.' in line or '_' in line or line.endswith(
                (".js.map", ".js", ".map", ".min.js", ".min.map", '.html')))
            and not (line.endswith(".js.map") or line.endswith(".js") or line.endswith(".map") or line.endswith(
                ".min.js") or line.endswith(".min.map")) or line.endswith(".min")]

        lines = [line for line in lines if
                 any(token.is_alpha and not token.is_stop and len(token.text) > 1 for token in nlp(line.lower()))]

        # Calculate changes
        new_size = len(lines) + len(second_lines)
        removed = original_size - new_size

        print(f"\n{YELLOW}[INFO]{RESET} Removed {GREEN}{removed}{RESET} lines.")
        print(f"\n{YELLOW}[INFO]{RESET} Wordlist is now {GREEN}{new_size}{RESET} lines.")
        
        if not os.path.exists(host):
            os.system(f"mkdir {host}")
            
        with open(f'{host}\{host}_wordlist.txt', 'w', encoding="utf-8") as f:

            for item in lines:
                f.write(f"{item}\n")

            for second_item in second_lines:
                f.write(f"{second_item}\n")

        print(f'\n{GREEN}[SUCCESS]{RESET} Wordlist saved to {BLUE}{host}\{host}_wordlist.txt{RESET}')

    except:
        print(
            f"\n{RED}[WARN]{RESET} Please install Spacy by issuing these commands in the command line:\n{RED}[WARN]{RESET} pip install spacy\n{RED}[WARN]{RESET} python -m spacy download en_core_web_sm")


def entropy(string):
    # "Calculates the Shannon entropy of a string"
    # get probability of chars in string
    prob = [float(string.count(c)) / len(string) for c in dict.fromkeys(list(string))]

    # calculate the entropy
    entropy = - sum([p * math.log(p) / math.log(2.0) for p in prob])

    return entropy


def wordlist_creator(file, host):
    tree = ET.parse(file)
    root = tree.getroot()
    wordlist = []

    print(f"\n{YELLOW}[INFO]{RESET} Please wait, it might take a few minutes.{RESET}")

    for i in root:

        # preserve subdomains, file/dir names with . - _
        wordlist += re.split('\/|\?|&|=', i[1].text)

        # get subdomain names and break up file names
        wordlist += re.split('\/|\?|&|=|_|-|\.|\+', i[1].text)

        # get words from cookies, headers, POST body requests
        wordlist += re.split('\/|\?|&|=|_|-|\.|\+|\:| |\n|\r|"|\'|<|>|{|}|\[|\]|`|~|\!|@|#|\$|;|,|\(|\)|\*|\|',
                             urllib.parse.unquote(base64.b64decode(i[8].text)))

        # response
        if i[12].text is not None:
            wordlist += re.split(
                '\/|\?|&|=|_|-|\.|\+|\:| |\n|\r|\t|"|\'|<|>|{|}|\[|\]|`|~|\!|@|#|\$|;|,|\(|\)|\*|\^|\\\\|\|',
                urllib.parse.unquote(base64.b64decode(i[12].text)))

    auxiliaryList = list(set(wordlist))
    final = []

    for word in auxiliaryList:
        if word.isalnum() or '-' in word or '-' in word or '_' in word:
            if len(word) == 2:
                continue

            en = entropy(word)
            # remove "random strings" that are high entropy
            if en < 4.4:
                final.append(word)

    final.sort()

    print(f"\n{YELLOW}[INFO]{RESET} Wordlist is {GREEN}{len(final)}{RESET} lines.")

    cleaning(host, final)


def avgEntropyByChar(en, length):
    # calulate "average" entropy level
    return en / length


def is_any_process_alive(processes):
    return True in [p.is_alive() for p in processes]


def adjust_column_widths_disk(filepath):
    workbook = load_workbook(filepath)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

    workbook.save(filepath)


def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.3
        sheet.column_dimensions[column].width = adjusted_width


def postMan(file):
    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()

    headers_list = []
    unique_path = set()
    unique_name_path = set()
    counter = 0

    for i in root:
        host = i.find('host').text
        break

    domain_output = i.find('host').text

    POST_Requests = []

    postman = {

        "info": {
            "_postman_id": "my-postman-id",
            "name": f"{host} API Endpoints",
            "description": "API Endpoints Documentation - Generated by Burp Collector",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "items": []
    }

    # Looping through each request/response
    for i in root:
        # Searching for responses only
        response = i.find('response').text
        status_code = i.find('status').text

        if status_code is not None:
            status_code = int(i.find('status').text)
        else:
            continue

        if response is None:
            continue
        # Decoding the response
        content = base64.b64decode(response)
        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: .*?(?:\r\n|\r|\n)', re.DOTALL)
        match = content_type_pattern.search(content)

        if match:
            content_type_header = match.group().decode('latin-1')

        method = i.find('method').text

        # Searching for responses only
        request = i.find('request').text
        # Decoding the response
        content_request = base64.b64decode(request)
        match_request = content_type_pattern.search(content_request)
        content_type_header_request = None

        # Decoding the request
        content_request = base64.b64decode(request)
        content_request = content_request.split(b'\r\n\r\n', 2)[0].decode('latin-1')
        # Split the string by newline and exclude the first line
        headers_lines = content_request.split('\n')[2:]

        # Join the remaining lines back together
        headers_lines = '\n'.join(headers_lines)

        headers_list = re.findall(r'(?P<name>.*?): (?P<value>.*?)(?:\r?\n|$)', headers_lines)

        headers_list = [{'key': key, 'value': value} for key, value in headers_list if value]

        host = i.find('host').text
        protocol = i.find('protocol').text
        url = i.find('url').text
        path = i.find('path').text
        domain_output = i.find('host').text
        success_codes = [200, 201, 202, 203, 204, 205, 206, 207, 208, 226]

        if match_request is not None:
            content_type_header_request = match_request.group(0).decode('latin-1')

        if method == 'POST' or method == 'DELETE' or method == 'PUT' or method == 'PATCH' or method == 'TRACE' or method == 'CONNECT':

            if match:
                content_type_header = match.group().decode('utf-8')
                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                if content_type_header is not None and 'application/json' in content_type_header or 'application/xml' in content_type_header and status_code in success_codes:

                    name_path = re.sub(r'\?.*', '', path)

                    if name_path in unique_name_path:
                        continue
                    else:
                        unique_name_path.add(name_path)

                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue

                    else:
                        path = path
                        if path.endswith("/"):
                            path = path[:-1]

                        content_request = base64.b64decode(request)
                        body = content_request.split(b'\r\n\r\n', 1)[1].decode('latin-1')

                        # create a dictionary to represent the POSTMAN collection
                        loop_collection = {
                            "name": name_path,
                            "request": {
                                "method": method,
                                "header": headers_list,
                                "url": {
                                    "raw": url,
                                    "host": [
                                        f"{protocol}://{host}"
                                    ],
                                    "path": [
                                        path
                                    ]
                                },
                                "description": "Not documented",
                                "body": {
                                    "mode": "raw",
                                    "raw": body
                                }
                            },
                            "response": []
                        }

                        POST_Requests.append(loop_collection)

                        if path not in unique_path and body != '[]' and content_type_header_request is not None:
                            unique_path.add(path)

                            # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                            if 'application/json' in content_type_header_request:
                                mode = "raw"
                                kv_pairs = body

                            elif 'application/xml' in content_type_header_request:
                                mode = "raw"
                                kv_pairs = body

                            elif 'application/x-www-form-urlencoded' in content_type_header_request:
                                mode = "urlencoded"
                                # URL decode the string
                                body = urllib.parse.unquote(body)

                                # Split the string by "&" to get a list of key-value pairs
                                pairs_list = body.split("&")

                                # Initialize an empty list to store the key-value pairs
                                kv_pairs = []

                                # Loop through each pair and split it by "=" to get the key and value separately
                                for pair in pairs_list:
                                    if "=" in pair:
                                        values = pair.split('=', 1)
                                        if values[1] == "":
                                            values.insert(1, f"<{values[0]}>")
                                        # Append the key-value pair to the kv_pairs list
                                        kv_pairs.append({"key": values[0], "value": values[1]})

                            else:
                                mode = "raw"
                                kv_pairs = body

                            # create a dictionary to represent the POSTMAN collection
                            loop_collection = {
                                "name": name_path,
                                "request": {
                                    "method": method,
                                    "header": headers_list,
                                    "url": {
                                        "raw": url,
                                        "host": [
                                            f"{protocol}://{host}"
                                        ],
                                        "path": [
                                            path
                                        ]
                                    },
                                    "description": "Not documented",
                                    "body": {
                                        "mode": mode,
                                        mode: kv_pairs
                                    }
                                },
                                "response": []
                            }

                            if str(loop_collection['request']['url']['path']) in str(POST_Requests):
                                continue

            mime_types = ['application/x-www-form-urlencoded', 'multipart/form-data', 'application/json',
                          'application/xml', 'text/plain']

            if match_request:

                for mime in mime_types:
                    if mime in content_type_header_request:

                        name_path = re.sub(r'\?.*', '', path)

                        if name_path in unique_name_path:
                            continue
                        else:
                            unique_name_path.add(name_path)

                        if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith(
                                "json?"):
                            continue

                        else:

                            if path.endswith("/"):
                                path = path[:-1]

                            # Decoding the request
                            content_request = base64.b64decode(request)

                            body = content_request.split(b'\r\n\r\n', 1)[1].decode('latin-1')

                            if path not in unique_path and body != '[]':
                                unique_path.add(path)

                                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                                if 'application/json' in content_type_header_request:
                                    mode = "raw"
                                    kv_pairs = body

                                elif 'application/xml' in content_type_header_request:
                                    mode = "raw"
                                    kv_pairs = body

                                elif 'application/x-www-form-urlencoded' in content_type_header_request:
                                    mode = "urlencoded"
                                    # URL decode the string
                                    body = urllib.parse.unquote(body)

                                    # Split the string by "&" to get a list of key-value pairs
                                    pairs_list = body.split("&")

                                    # Initialize an empty list to store the key-value pairs
                                    kv_pairs = []

                                    # Loop through each pair and split it by "=" to get the key and value separately
                                    for pair in pairs_list:
                                        if "=" in pair:
                                            values = pair.split('=', 1)
                                            if values[1] == "":
                                                values.insert(1, f"<{values[0]}>")
                                            # Append the key-value pair to the kv_pairs list
                                            kv_pairs.append({"key": values[0], "value": values[1]})

                                else:
                                    mode = "raw"
                                    kv_pairs = body

                                # create a dictionary to represent the POSTMAN collection
                                loop_collection = {
                                    "name": name_path,
                                    "request": {
                                        "method": method,
                                        "header": headers_list,
                                        "url": {
                                            "raw": url,
                                            "host": [
                                                f"{protocol}://{host}"
                                            ],
                                            "path": [
                                                path
                                            ]
                                        },
                                        "description": "Not documented",
                                        "body": {
                                            "mode": mode,
                                            mode: kv_pairs
                                        }
                                    },
                                    "response": []
                                }

                                POST_Requests.append(loop_collection)

        if method == 'GET':

            if match:
                content_type_header = match.group().decode('utf-8')
                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                if 'application/json' in content_type_header or 'application/xml' in content_type_header and status_code in success_codes:

                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue

                    else:
                        second_path = path = i.find('path').text
                        name_path = re.sub(r'\?.*', '', path)

                        if name_path in unique_name_path:
                            continue
                        else:
                            unique_name_path.add(name_path)

                        if "?" in path:
                            path = path.split("?")
                            query_parameters = path[1]
                            path = path[0]
                        if "?" not in path:
                            path = path

                        if path in unique_path:
                            continue
                        else:
                            unique_path.add(path)
                            counter += 1

                        # create a dictionary to represent the POSTMAN collection
                        loop_collection = {
                            "name": f"{name_path}",
                            "request": {
                                "method": "GET",
                                "header": headers_list,
                                "url": {
                                    "raw": url,
                                    "host": [
                                        f"{protocol}://{host}"
                                    ],
                                    "path": [
                                        path
                                    ]
                                },

                                "query": [],
                            },
                            "response": []
                        }

                        if "?" in second_path:

                            # URL decode the string
                            parameters = urllib.parse.unquote(query_parameters)
                            # Split the string by "&" to get a list of key-value pairs
                            pairs_list = parameters.split("&")

                            # Initialize an empty list to store the key-value pairs
                            kv_pairs = []

                            # Loop through each pair and split it by "=" to get the key and value separately
                            for pair in pairs_list:
                                if "=" in pair:
                                    values = pair.split('=', 1)
                                    if values[1] == "":
                                        values.insert(1, f"<{values[0]}>")

                                    # Append the key-value pair to the kv_pairs list
                                    kv_pairs.append({"key": values[0], "value": values[1]})

                            loop_collection["request"]["url"]["query"] = kv_pairs

                        postman["items"].append(loop_collection)

    for post in POST_Requests:
        counter += 1
        postman["items"].append(post)

    postman['info'].update({"description": f"Total {counter} API Endpoints - Generated by Burp Collector"})

    if counter == 0:
        print(f'\n{RED}[WARN]{RESET} No API Endpoints found in order to generate JSON file for Postman.')

    if counter > 0:
        # convert the dictionary to a JSON string and print it
        json_collection = json.dumps(postman)

        if not os.path.exists(host):
            os.system(f"mkdir {host}")

        with open(f"{host}\Postman_Collection.json", "w") as f:
            f.write(f"{json_collection}")
            print(
                f'\n{GREEN}[SUCCESS]{RESET} {ORANGE}Postman_Collection.json{RESET} created in {BLUE}{domain_output}{RESET} directory.')
            print(f'\n{YELLOW}[INFO]{RESET} You can open it with {ORANGE}Postman!{RESET}')

        # Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)


def js_file(file, wb, sheet):
    js_list = []

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()

    for i in root:

        domain = i.find('host').text

        url = i.find('url').text

        if url.endswith('.js') or 'js?' in url or url.endswith('.map') or 'map?' in url:
            print(f'\n{YELLOW}Testing:{RESET} {url}')
            if 'js?' in url or 'map?' in url:
                if url.endswith('/') or url.endswith('\\'):
                    url = url[:-1]
                url = url.split('?')[0]

            js_list.append(url)

    data = []

    new_list = []

    for js_file in js_list:

        if js_file not in new_list:
            new_list.append(js_file)

        elif js_file in new_list:
            continue

    for js in new_list:
        data.append([js])

    for row in data:
        sheet.append(row)
        # Setting font of the cell to Calibri 14
        row = sheet.max_row
        sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)

    if not os.path.exists(domain):
        os.system(f"mkdir {domain}")

    if len(data) != 0:
        # Adjust column widths
        adjust_column_widths(sheet)
        wb.save(f'{domain}\JS_Files.xlsx')

        print(f'\n{GREEN}[SUCCESS]{RESET} {ORANGE}JS_Files.xlsx{RESET} created in {BLUE}{domain}{RESET} directory.')


# Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)
def json_file(file, wb):
    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()

    data = []

    # Looping through each request/response
    for i in root:
        # Searching for responses only
        response = i.find('response').text
        url = i.find('url').text
        if response is None:
            continue
        # Decoding the response
        content = base64.b64decode(response)
        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: .*?(?:\r\n|\r|\n)', re.DOTALL)
        match = content_type_pattern.search(content)

        if match:
            content_type_header = match.group().decode('utf-8')
            # if Content-Type is equals to JSON/XML it will be added to the xlsx file
            if 'application/json' in content_type_header:

                path = i.find('path').text
                domain = i.find('host').text

                if path.endswith(".json") or path.endswith(".json?"):

                    data.append([domain, path, url])

                elif ".json?" in path and not path.endswith(".json?"):

                    split_path = path.split("?")[0]
                    split_url = url.split("?")[0]
                    data.append([domain, split_path, split_url])

                else:
                    continue

    if len(data) > 0:
        # Sorting and removing duplicates
        data = sorted(list(set([tuple(row) for row in data])))

        if args.directory:

            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_json(host, "JSON Files")

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_json(host, "JSON Files")

        for row in data:
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            # Removing duplicates and sorting rows

            if not os.path.exists(domain):
                os.system(f"mkdir {domain}")

            # Adjust column widths
            adjust_column_widths(sheet)
            wb.save(f'{domain}\JSON_Files.xlsx')

        print(f'\n{GREEN}[SUCCESS]{RESET} {ORANGE}JSON_Files.xlsx{RESET} created in {BLUE}{domain}{RESET} directory.')

    else:
        print(f'\n{RED}[WARN]{RESET} No JSON Files found.')

    # Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)


def api_collector(file):
    
    def is_exact_match(data, search_pattern):
        for item in data:
            if isinstance(item, list):
                if is_exact_match(item, search_pattern):
                    return True
            else:
                if item == search_pattern:
                    return True
        return False

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()
    data = []
    new_data = []
    success_codes = [200, 201, 202, 203, 204, 205, 206, 207, 208, 226]

    if args.status_code:
        status_code_list = args.status_code

        if "," in status_code_list:
            status_code_split = status_code_list.split(",")
            for sc in status_code_split:
                sc = int(sc)
                success_codes.append(sc)
        else:

            success_codes.append(int(status_code_list))

    valid_content_types = [
    'application/json',
    'application/xml',
    'text/xml',
    'xml', 
    'application/soap',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]

    def replace_variables_in_path(path):

        final_words = []

        id_matches = re.findall(r'/\d+[a-zA-Z0-9]*', path)
        if len(id_matches) == 0:
            id_matches = re.findall(r'/(?:(\d+[a-zA-Z0-9]*)/)+', path)

        ids = [match for match in id_matches if not match.isalpha()]

        matches = re.findall(r'\/([\w-]+)(?=\/\d+)', path)

        variables = []
        for variable in matches:
            variables.append(variable)

        for id in matches:
            if id.endswith("s"):
                id = id[:-1]
            final_words.append("{" + id + "Id" + "}")

        final_path = path

        for id, final_word in zip(ids, final_words):
            final_path = re.sub(r'\b' + re.escape(id) + r'\b', '/' + final_word, final_path)

        return final_path


    # Looping through each request/response
    for i in root:
        flag = False
        # Searching for responses only
        response = i.find('response').text
        if response is None:
            continue
  
        # Decoding the response
        content = base64.b64decode(response)

        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: ([^\r\n]+)', re.IGNORECASE)

        match_response = content_type_pattern.search(content)

        # Searching for responses only
        request = i.find('request').text
        status_code = i.find('status').text
        if status_code is not None:
            status_code = int(i.find('status').text)
        else:
            continue

        if request is None:
            continue

        # Decoding the response
        content_request = base64.b64decode(request)

        match_request = content_type_pattern.search(content_request)

        if match_request is not None:
            content_type_header_request = match_request.group(0).decode('latin-1')

        # Based on responses
        if match_response:
            content_type_header = match_response.group().decode('utf-8')

            if any(content in content_type_header for content in valid_content_types) and status_code in success_codes:

                path = i.find('path').text
                
                # if Content-Type is equals to JSON/XML it will be added to the xlsx file

                if status_code in success_codes:
                    
                    path = i.find('path').text
                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue
                    

                    path = path.split("?")[0]
                    method = i.find('method').text
                    domain = i.find('host').text
                    pathCheck = path.split("/")
                    tested = "       v        "

                    temp_path = replace_variables_in_path(path)

                    if temp_path is not None:

                        path = temp_path
                        
                    for j in range(len(pathCheck)):
                        if pathCheck[j].count("-") >= 3:
                            pathCheck[j] = "{UUID}"
                            path = "/".join(pathCheck)
                            
                    for first_uuid in data:
                        if path.endswith("{UUID}") and first_uuid[1].endswith("{UUID}"):
                            temp_path = path.replace("/{UUID}", "")
                            first_uuid = first_uuid[1].replace("/{UUID}", "")
                            if temp_path == first_uuid:
                                flag = True
                                continue

                    # uniq_method = method
                    # for uniq_endpoint in data:
                    #     if len(uniq_endpoint) >= 3:
                    #         inner_method = str(uniq_endpoint[2])
                    #         if path == uniq_endpoint[1] and uniq_method not in inner_method:
                    #             updated_value = uniq_method + "/" + inner_method
                    #             uniq_endpoint[2] = updated_value
                    #             continue

                    # if is_exact_match(data, path):
                    #     continue
                    # else:
                    #     path = path.replace("//", "/")
                    #     data.append([domain, path, method, tested])
                    path = path.replace("//", "/")
                    data.append([domain, path, method, tested])


        # Based on requests
        mime_types = ['application/x-www-form-urlencoded', 'multipart/form-data', 'application/json', 'application/xml',
                      'text/plain', 'text/xml', "application/json; charset=utf-8"]

        if match_request:

            method = i.find('method').text
            if status_code in success_codes:

                for mime in mime_types:
                    flag = False
                    
                    if mime in content_type_header_request:

                        path = i.find('path').text
                        path = path.split("?")[0]
                        tested = "       v        "
                        unique_path = path.split("/")
                        path_start = unique_path[1]
                        path_last = unique_path[-1]
                        path_start = "/" + str(path_start)
                        path_last = "/" + str(path_last)
                        pathCheck = path.split("/")
                        ids = []

                        if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                            continue

                        method = i.find('method').text
                        domain = i.find('host').text

                        temp_path = replace_variables_in_path(path)
                        if temp_path is not None:
                            path = temp_path

                        # uniq_method = method
                        # for uniq_endpoint in data:
                        #     if len(uniq_endpoint) >= 3:
                        #         inner_method = str(uniq_endpoint[2])
                        #         if path == uniq_endpoint[1] and uniq_method not in inner_method:
                        #             updated_value = uniq_method + "/" + inner_method
                        #             uniq_endpoint[2] = updated_value
                        #             continue

                        for j in range(len(pathCheck)):
                            if pathCheck[j].count("-") >= 3:
                                pathCheck[j] = "{UUID}"
                                path = "/".join(pathCheck)
                        
                        for second_uuid in data:
                            if path.endswith("{UUID}") and second_uuid[1].endswith("{UUID}"):
                                temp_path = path.replace("/{UUID}", "")
                                second_uuid = second_uuid[1].replace("/{UUID}", "")
                                if temp_path == second_uuid:
                                    flag = True
                                    continue
                        if flag:
                            continue

                        # if is_exact_match(data, path):
                        #     continue
                        # else:
                        #     path = path.replace("//", "/")
                        #     data.append([domain, path, method, tested])
                        path = path.replace("//", "/")
                        data.append([domain, path, method, tested])

    if len(data) > 0:

        # Sorting and removing duplicates
        data = sorted(list(set([tuple(row) for row in data])))

        if args.directory:

            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_api_collector(host, "API_Endpoints")

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_api_collector(host, "API_Endpoints")

        # Sorting and removing duplicates
        for row in data:
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=4).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
            # Removing duplicates and sorting rows

        # Sorting and removing duplicates
        for row in new_data:
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=4).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
            # Removing duplicates and sorting rows

        if not os.path.exists(domain):
            os.system(f"mkdir {domain}")

            # Adjust column widths
        adjust_column_widths(sheet)
        wb.save(f'{domain}\API_Endpoints.xlsx')
        print(
            f'\n{GREEN}[SUCCESS]{RESET} {ORANGE}API_Endpoints.xlsx{RESET} created in {BLUE}{domain}{RESET} directory.')


    else:
        print(f'\n{RED}[WARN]{RESET} No API Endpoints found (XML/JSON Content-Type) with APIs Collector feature.')


def parse_args():
    banner = f"""                                                                                                                                           
██████╗ ██╗   ██╗██████╗ ██████╗      ██████╗ ██████╗ ██╗     ██╗     ███████╗ ██████╗████████╗ ██████╗ ██████╗               
██╔══██╗██║   ██║██╔══██╗██╔══██╗    ██╔════╝██╔═══██╗██║     ██║     ██╔════╝██╔════╝╚══██╔══╝██╔═══██╗██╔══██╗     
██████╔╝██║   ██║██████╔╝██████╔╝    ██║     ██║   ██║██║     ██║     █████╗  ██║        ██║   ██║   ██║██████╔╝  
{BLUE}██╔══██╗██║   ██║██╔══██╗██╔═══╝     ██║     ██║   ██║██║     ██║     ██╔══╝  ██║        ██║   ██║   ██║██╔══██╗
{BLUE}██████╔╝╚██████╔╝██║  ██║██║         ╚██████╗╚██████╔╝███████╗███████╗███████╗╚██████╗   ██║   ╚██████╔╝██║  ██║
{BLUE}╚═════╝  ╚═════╝ ╚═╝  ╚═╝╚═╝          ╚═════╝ ╚═════╝ ╚══════╝╚══════╝╚══════╝ ╚═════╝   ╚═╝    ╚═════╝ ╚═╝  ╚═╝ 
                                                                                                                 
{RESET}Pentesters & Bug Hunters Swiss Army Knife 

{BLUE}Developed by Sagiv Michael{RESET}
                                                                                                                          """
    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter, description=banner)
    parser.add_argument('-f', '--file', type=str, required=False,
                        help='Burp File (Right Click on the domain in the Target Scope and select save selected items and select Base64 encode).')
    parser.add_argument('-dr', '--directory', type=str, required=False,
                        help='Directroy containing all Burp Suite output files.')
    parser.add_argument('-a', '--all', required=False, action="store_true",
                        help='Use all methods below - Can be be slow depends on the size of the project, so leave it running in the background.')
    parser.add_argument('-A', '--api', required=False, action="store_true",
                        help='Collect and extract all API Endpoints you interacted with during the test (SOAP, REST, and GraphQL) to an Excel file - Highly recommended')
    parser.add_argument('--status_code', required=False, type=str, default=False,
                        help='Status codes to include in the --api feature. Example: 401,402 (default 20x).')
    parser.add_argument('-p', '--postman', required=False, action="store_true",
                        help='Collect and extract all API Endpoints with their body and parameters to a Postman collection - Highly recommended.')
    parser.add_argument('-w', '--wordlist', required=False, action="store_true",
                        help="Create a tailored wordlist for your target (Based on Requests/Responses/Cookies/Headers etc') - Recommended!")
    parser.add_argument('-m', '--map', required=False, action="store_true",
                        help="Convert map files to their original Javascript source code.")
    parser.add_argument('-D', '--dependency', required=False, action="store_true",
                        help="Convert map files to their original Javascript source code and check if the dependencies exists in npmjs.com.")
    parser.add_argument('-P', '--paramspider', required=False, action="store_true",
                        help="Collecting URLs with parameters and dumping them to a file with a FUZZ keyword.")
    parser.add_argument('-d', '--domain', required=False, action="store_true",
                        help='Collect and extract all subdomains encountered during the test into an Excel file - Fast.')
    parser.add_argument('-j', '--json', required=False, action="store_true",
                        help='Collect JSON files based on Burp response via REGEX to Excel file - Fast.')
    parser.add_argument('-J', '--js', required=False, action="store_true",
                        help='Collect and extract all JS/MAP URLs encountered during the test to an Excel file - Fast.')
    parser.add_argument('-pe', '--postoexcel', required=False, action="store_true",
                        help='Convert Postman collections to an Excel file - Recommended!')
    parser.add_argument('-i', '--path', required=False, action="store_true",
                        help='Collect and extract possible APIs found in files during the test to an Excel file - It might generate a lot of junk, but it could be helpful if used right.')
    parser.add_argument('-s', '--secrets', required=False, action="store_true",
                        help="Collect and extract all possible secrets (AWS/Google/Firebase, etc') that might be disclosed - Most of the time the output will be False-Positive.")
    parser.add_argument('-u', '--urls', required=False, action="store_true",
                        help='Collect and extract all URLs encountered during the test to an Excel file - This can be slow depending on the project size.')
    parser.add_argument('-t', '--threads', required=False, type=int, default=os.cpu_count(),
                        help='Number of processes to run in parallel (Default is the number of your CPU cores).')
    parser.add_argument('-v', '--verbose', required=False, action="store_true",
                        help='If set, output will be printed to the screen with colors.')
    parser.add_argument("--update", action="store_true", help="Update to the latest version.")
    parser.add_argument("--version", action="store_true", help="Version check.")

    return parser.parse_args()


def create_worksheet(host, wb):
    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = host

    sheet['A1'] = 'URL Tested'
    sheet['B1'] = 'Regex'
    sheet['C1'] = 'Matched Pattern'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet, wb


def create_worksheet_main(string):
    # Creating a new Workbook object
    wb = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = string

    sheet['A1'] = 'URL Tested'
    sheet['B1'] = 'Regex'
    sheet['C1'] = 'Matched Pattern'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet, wb


def create_worksheet_api_collector(host, string):
    # Creating a new Workbook object
    wb = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = f"{host}_{string}"

    # Writing data to sheet
    sheet = wb.active
    sheet.title = "API Endpoints"
    sheet['A1'] = 'HOST'
    sheet['B1'] = 'ENDPOINT'
    sheet['C1'] = 'METHOD'
    sheet['D1'] = 'TESTED?'
    header_font = Font(name='Calibri', size=20, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    sheet['A1'].font = header_font
    sheet['A1'].alignment = align_center

    sheet['B1'].font = header_font
    sheet['B1'].alignment = align_center

    sheet['C1'].font = header_font
    sheet['C1'].alignment = align_center

    sheet['D1'].font = header_font
    sheet['D1'].alignment = align_center

    return sheet, wb


def create_worksheet_json(host, string):
    # Creating a new Workbook object
    wb = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = f"{host}_{string}"

    # Writing data to sheet
    sheet = wb.active
    sheet.title = "JSON Files"
    sheet['A1'] = 'HOST'
    sheet['B1'] = 'JSON File'
    sheet['C1'] = 'URL'

    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet, wb


def create_worksheet_js(string):
    # Creating a new Workbook object
    wb = Workbook()

    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = string

    # Writing data to sheet
    sheet = wb.active
    sheet.title = "JS Files"
    sheet['A1'] = 'JS/MAP File'

    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font

    return sheet, wb


def postmanToJSON(json_directory, data):
    for filename in os.listdir(json_directory):
        if filename.endswith('.json'):
            file_path = os.path.join(json_directory, filename)

            # Open and load the JSON file
            with open(file_path, 'r', encoding='utf-8') as file:
                postman_data = json.load(file)

            extract_endpoints(postman_data)
            # Sorting and removing duplicates
            wb, sheet = create_worksheet_postmantoexcel()

            for row in data:
                sheet.append(row)
                # Setting font of the cell to Calibri 14
                row = sheet.max_row
                sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
                sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
                sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)

    adjust_column_widths(sheet)
    wb.save(f'API_Endpoints.xlsx')
    print(f"{GREEN}[SUCCESS]{RESET} {ORANGE}API_Endpoints.xlsx{RESET} created in {BLUE}{os.getcwd()}{RESET} directory.")


def match(regex, content, url, host, sheet, wb, matched_patterns, string, args, final_xlsx):
    data = []

    try:
        # Iterating over the given regex patterns to search for matches
        for key, value in regex.items():
            if args.verbose:
                # Printing the URL being tested
                print(f'{YELLOW}[INFO]{RESET} Testing Regex: {BLUE}{key}: {value}{RESET}')
                print(f'{YELLOW}[INFO]{RESET} Testing URL: {BLUE}{url}{RESET}')

            # Searching for a match of the current regex pattern
            pattern = re.compile(value, re.IGNORECASE | re.MULTILINE)

            # import jsbeautifier
            # print(f'\n{BLUE}[+] Beautifying the code, please wait.{RESET}\n')
            # content = jsbeautifier.beautify(content)
            # print(content)
            match = re.findall(pattern, content)

            if match:
                flag = False
                for matched_pattern in match:

                    # Only add unique matched patterns to the set
                    if matched_pattern not in matched_patterns:
                        if string == 'Path_and_Endpoints':
                            parts = matched_pattern.split("/")
                            if "." in parts[-1]:
                                continue

                            threshold = 4
                            entropy_string = str(matched_pattern)[1:]  # exclude the first forward slash
                            # Calculate the entropy of the string
                            prob = [float(entropy_string.count(c)) / len(entropy_string) for c in
                                    dict.fromkeys(list(entropy_string))]
                            # calculate the entropy
                            entropy = - sum([p * math.log(p) / math.log(2.0) for p in prob])

                            try:
                                flag = False
                                if entropy < threshold:

                                    if "//" in str(matched_pattern):
                                        continue

                                    if "/" in str(matched_pattern):
                                        if len(matched_pattern) == 2 or len(matched_pattern) == 3:
                                            continue

                                    if "\/\/" in str(matched_pattern):
                                        continue

                                    if '\/\/' in str(matched_pattern):
                                        continue

                                    if "//" in str(matched_pattern):
                                        continue

                                    if flag:
                                        continue

                                    if "-" in matched_pattern or "_" in matched_pattern:
                                        continue

                                    if flag:
                                        continue

                                if entropy > threshold:
                                    continue

                            except Exception as error:
                                print(error)
                                traceback.print_exc()

                        matched_patterns.add(matched_pattern)
                        matched_pattern = list(matched_pattern)

                        if len(matched_pattern) != 0:
                            if len(matched_pattern) > 2:
                                if matched_pattern[2].endswith('\\'):
                                    matched_pattern[2] = matched_pattern[2][:-1]
                                if '\\/' in matched_pattern[2]:
                                    matched_pattern[2] = matched_pattern[2].replace('\\/', '/')
                            if 'http' == matched_pattern[0]:
                                matched_pattern[0] = 'http://'

                            elif 'https' == matched_pattern[0]:
                                matched_pattern[0] = 'https://'

                            elif 'ftp' == matched_pattern[0]:
                                matched_pattern[0] = 'ftp://'

                            matched_pattern = "".join(matched_pattern)
                            if matched_pattern.endswith("?"):
                                matched_pattern = matched_pattern[:-1]

                            if args.verbose:
                                print(
                                    f'{YELLOW}[INFO]{RESET} Matched regex: {GREEN}{key}: {value}{RESET} with pattern: {GREEN}{matched_pattern}{RESET}')

                            if '\\/' in matched_pattern:
                                matched_pattern = matched_pattern.replace('\\/', '/')

                            # Only append unique regex matches to data
                            if matched_pattern not in [row[2] for row in data]:
                                if host in url:
                                    data.append([url, f"{key}: ({value})", matched_pattern])

            elif not match:
                if args.verbose:
                    print(f'\n{RED}[WARN]{RESET} No match has been found for this regex.')

        # Removing duplicates and sorting rows
        data = sorted(list(set([tuple(row) for row in data])))

        for row in sorted(data):
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)

            if not os.path.exists(host):
                os.system(f"mkdir {host}")

            final_xlsx.append(f"{host}\{string}.xlsx")
            wb.save(f"{host}\{string}.xlsx")
            wb.close()

        return host

    except Exception as error:
        # print the error message and traceback
        print(error)
        traceback.print_exc()
        exit(1)


def main(file, tool_method, sheet, wb, uri_finder, regex_secrets, api_extractor, args, matched_patterns, final_xlsx):
    # Create an XML Object
    tree = ET.parse(file)
    root = tree.getroot()
    endpoint_check = ""
    final_host = None
    empty_list = []
    workbooks = {}
    empty_url_list = []
    flag = False

    # Checks if file imoprted from HTTP History to determine different hosts

    for i in root:

        response = i.find('response').text

        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        hostOne = i.find('host').text
        break

    counter = 0
    for i in root:
        counter += 1
        response = i.find('response').text

        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        hostTwo = i.find('host').text

        if hostTwo != hostOne:
            flag = True
            break

        elif counter > 15:
            break

    workbooks = {}
    for i in root:
        counter += 1
        response = i.find('response').text

        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        unique_host = i.find('host').text

        if flag:
            if unique_host not in workbooks:
                wb = Workbook()
                workbooks[unique_host] = wb
            else:
                wb = workbooks[unique_host]

            if tool_method == "Secrets":
                sheet, wb = create_worksheet(unique_host, wb)
            elif tool_method == "Path_and_Endpoints":
                sheet, wb = create_worksheet(unique_host, wb)
            elif tool_method == "URLs":
                sheet, wb = create_worksheet(unique_host, wb)
            elif tool_method == "Sub-Domains":
                sheet, wb = create_worksheet(unique_host, wb)

        url = i.find('url').text

        if url not in empty_url_list:
            empty_url_list.append(url)

        elif url in empty_url_list:
            continue

        unique_path = i.find('path').text

        if unique_path not in empty_list:
            empty_list.append(unique_path)

        elif unique_path in empty_list:
            continue

        parsed_url = urlparse(url)
        domain_parts = parsed_url.netloc.split(".")
        if len(domain_parts) > 2:
            domain = ".".join(domain_parts[1:])
        else:
            domain = parsed_url.netloc

        # Regex for finding subdomains
        sub_domains = {

            'Sub-Domain': f'[^0-9_\-][0-9A-Za-z-_]+\.{domain}',
            'Sub-Domain': f'https?://[a-zA-Z0-9\.]+\.{domain}'
        }

        if tool_method == "Path_and_Endpoints":

            print(f'\n{YELLOW}Testing:{RESET} {url}')

            endpoint_check += "good"
            host = match(api_extractor, content, url, unique_host, sheet, wb, matched_patterns, tool_method, args,
                         final_xlsx)
            if host is not None:
                final_host = host

        if tool_method == "URLs":
            print(f'\nTesting URL: {url}')
            host = match(uri_finder, content, url, unique_host, sheet, wb, matched_patterns, tool_method, args,
                         final_xlsx)

            if host is not None:
                final_host = host

        if tool_method == "Sub-Domains":

            print(f'\n{YELLOW}Testing:{RESET} {url}')
            host = match(sub_domains, content, url, unique_host, sheet, wb, matched_patterns, tool_method, args,
                         final_xlsx)
            if host is not None:
                final_host = host

        if tool_method == "Secrets":
            print(f'\n{YELLOW}Testing:{RESET} {url}')
            host = match(regex_secrets, content, url, unique_host, sheet, wb, matched_patterns, tool_method, args,
                         final_xlsx)
            if host is not None:
                final_host = host

    if final_host is not None:
        filename = f'{final_host}\{tool_method}.xlsx'

        if os.path.isfile(filename):
            print(
                f'\n{GREEN}[SUCCESS]{RESET} {ORANGE}{tool_method}.xlsx{RESET} created in {BLUE}{final_host}{RESET} directory.')

        elif endpoint_check == "":
            print(f'\n{RED}[WARN]{RESET} Nothing found for {tool_method} in {final_host}')
        else:
            print(f'\n{RED}[WARN]{RESET} Nothing found for {tool_method} in {final_host}')

    elif final_host is None:
        if endpoint_check == "":
            print(f'\n{RED}[WARN]{RESET} Nothing found for {tool_method} in {unique_host}')
        else:
            print(f'\n{RED}[WARN]{RESET} Nothing found for {tool_method} in {unique_host}')


def get_current_version():
    return "v1.1.5-offical"


def create_delete_script(script_path, python_script_path, zip_file_name):
    delete_script_path = os.path.join(script_path, "delete_script.bat")
    temp_delete_script_path = os.path.join(script_path, "temp_delete_script.bat")

    # Create the delete script
    with open(delete_script_path, "w") as delete_script:
        delete_script.write('@echo off\n')
        delete_script.write(f'del "{python_script_path}"\n')

        # Create a temporary batch script to delete the original batch script
        delete_script.write(f'echo @echo off > "{temp_delete_script_path}"\n')
        delete_script.write(f'echo del "{delete_script_path}" >> "{temp_delete_script_path}"\n')
        delete_script.write(f'timeout 4 >> "{temp_delete_script_path}"\n')  # Delay for about 5 seconds
        delete_script.write(f'echo del "{zip_file_name}" >> "{temp_delete_script_path}"\n')  # Delete temp script itself
        delete_script.write(f'echo del "{temp_delete_script_path}" >> "{temp_delete_script_path}"\n')  # Delete temp script itself

        # Execute the temporary batch script
        delete_script.write(f'start cmd /c "{temp_delete_script_path}"\n')

    return delete_script_path


def update_notifier():
    repository = 'https://api.github.com/repos/sAjibuu/Burp_Collector/releases/latest'.format(owner='owner_name',
                                                                                              repo='repo_name')
    session = requests.Session()
    response = session.get(repository)
    current_version = get_current_version()

    if response.status_code == 200:
        latest_version = response.json()['tag_name']
        if latest_version != current_version:
            print(
                f"\n{YELLOW}[INFO]{RESET} A new {GREEN}version{RESET} of Burp Collector is available to download! 🎉🎉🎉")


def check_for_updates():
    repository = 'https://api.github.com/repos/sAjibuu/Burp_Collector/releases/latest'.format(owner='owner_name',
                                                                                              repo='repo_name')
    session = requests.Session()
    response = session.get(repository)
    current_version = get_current_version()

    if response.status_code == 200:
        latest_version = response.json()['tag_name']
        if latest_version != current_version:
            print(f"\n{YELLOW}[INFO]{RESET} Downloading the latest version.")
            assets_url = response.json()['assets_url']
            assets_response = session.get(assets_url)
            if assets_response.status_code == 200:
                assets_json = assets_response.json()
                for asset in assets_json:
                    if asset['name'].endswith('.zip'):
                        asset_url = asset['url']
                        headers = {
                            "Accept": "application/octet-stream"
                        }
                        asset_response = session.get(asset_url, headers=headers)
                        if asset_response.status_code == 200:
                            package_name = asset['name']
                            description = response.json()["body"]
                            with open(asset['name'], 'wb') as f:
                                f.write(asset_response.content)

                                import platform
                                import zipfile

                                if "windows" in platform.system().lower():
                                    zip_file_name = package_name
                                    new_tool_name = zip_file_name.replace(".zip", ".py")
                                    # Open the zip file
                                    with zipfile.ZipFile(zip_file_name, 'r') as zip_ref:
                                        # Extract all contents to the current folder
                                        zip_ref.extractall()

                                    # Close the zip file handle
                                    zip_ref.close()

                                    # Get the path of the current script
                                    python_script_path = os.path.abspath(__file__)
                                    current_script_path = os.getcwd()
                                    # Create the delete script
                                    delete_script_path = create_delete_script(current_script_path, python_script_path, zip_file_name)

                                    print(f"\n{YELLOW}[INFO]{RESET} Download complete.")
                                    print(f"\n{YELLOW}[INFO]{RESET} Upgraded to: {latest_version}")
                                    print(
                                        f"\n{YELLOW}[INFO]{RESET} Burp Collector is extracted and is available to use as: {new_tool_name}")

                                    print("\nRelease Notes:")
                                    print(f"\n{YELLOW}{description}{RESET}\n")

                                    # Execute the delete script using subprocess
                                    import subprocess

                                    subprocess.call(delete_script_path, shell=True)

                                    # Terminate the script execution
                                    sys.exit()

                                elif "linux" in platform.system().lower():

                                    from pathlib import Path
                                    # Get the path of the current script
                                    current_script_path = os.path.abspath(__file__)

                                    os.system(f"mv ./{package_name} /tmp/{package_name}")
                                    os.system(f"unzip -q /tmp/{package_name} -d /tmp/")
                                    os.system(
                                        f"rsync --force --remove-source-files /tmp/Burp_Extractor_{latest_version}.py ./")
                                    os.system(f"rm -rf /tmp/{package_name}")
                                    os.system(f"rm -rf /tmp/requirements.txt")
                                    os.system(f"rm -rf {current_script_path}")

                                    print("Download complete.")
                                    print(f"Upgraded to: {latest_version}")
                                    print("Release Notes:\n" + description)
                                    sys.exit()

        else:
            print(f'\n{YELLOW}[INFO]{RESET} Burp Collector is on the {GREEN}latest{RESET} version!')
    else:
        print('Failed to retrieve the latest version from Github.')


def get_source_map(url, content):
    print("Retrieving Sourcemap from {}.".format(url))
    m = {}

    body = content

    # Unmarshal the body into the dictionary.
    print("Read {} bytes, parsing JSON.".format(len(body)))

    if body is not None and len(body) > 0:
        try:
            m = json.loads(body)
        except json.JSONDecodeError as e:
            print("Error decoding JSON:", str(e))
    else:
        print("Empty body received, skipping.")

    return m


def write_file(output_dir, content):
    if content is None:
        print("Content is None, skipping file write.")
        return

    os.makedirs(os.path.dirname(output_dir), exist_ok=True)
    with open(output_dir, 'w', encoding='utf-8') as f:
        f.write(content)
    print("Writing {} bytes to {}.".format(len(content), output_dir))


def clean_windows(p):
    m1 = re.compile(r'[?%*|:"<>]')
    return m1.sub('-', p)


def mapConverter(file):
    unique_js = []

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()
    success_codes = [200, 201, 202, 203, 204, 205, 206, 207, 208, 226, 401, 403]

    for i in root:
        url = i.find('url').text
        domain = i.find('host').text

        status_code = i.find('status').text
        if status_code is not None:
            status_code = int(i.find('status').text)
        else:
            continue

        if 'map?' in url:
            if url.endswith('/') or url.endswith('\\'):
                url = url[:-1]
            url = url.split('?')[0]

        if url.endswith('.map') and "css.map" not in url and status_code in success_codes:

            if url not in unique_js:
                unique_js.append(url)
            elif url in unique_js:
                continue

            response = i.find('response').text

            if response is None:
                continue

            content_response = base64.b64decode(response)
            content_response = content_response.decode('latin-1')

            try:
                data = content_response.split('\r\n\r\n', 1)
                data = data[1]
            except IndexError:
                data = content_response.split('\n\n', 1)
                data = data[1]

            if response is None:
                continue

            sm = get_source_map(url, data)

            if sm is None or not isinstance(sm, dict):
                print("Invalid sourcemap received, skipping.")
                continue

            print("Retrieved Sourcemap with version {}, containing {} entries.".format(sm.get('version'),
                                                                                       len(sm.get('sources', []))))

            sources = sm.get('sources')
            sources_content = sm.get('sourcesContent')

            if not sources:
                print("No sources found.")
                continue

            if not sources_content:
                print("No source content found.")
                continue

            if sm.get('version') != 3:
                print("[!] Sourcemap is not version 3. This is untested!")

            current_directory = os.getcwd()
            concatenated_path = os.path.join(current_directory, domain)

            os.makedirs(concatenated_path, exist_ok=True)

            for i, source_path in enumerate(sources):
                source_path = source_path.lstrip('/')
                source_path = source_path.replace("..", "").replace("/", "\\")
                source_path = source_path.replace("^", "").replace(".", "").replace("-", "").replace(" ", "")
                if os.name == 'nt':
                    source_path = clean_windows(source_path)

                script_path = concatenated_path + "\\" + source_path
                script_path = script_path.replace("\\\\", "\\")

                if i < len(sources_content):
                    script_data = sources_content[i]
                    write_file(script_path, script_data)

                else:
                    print("Missing source content for source path:", source_path)

            print(
                f"{GREEN}[SUCCESS]{RESET} The JavaScript source code extracted to {BLUE}{concatenated_path}{RESET} directory.")


def dependencyCheck(file, jsCheck=None):

    success_codes = [200, 201, 202, 203, 204, 205, 206, 207, 208, 226]

    if jsCheck is None:
        unique_js = []

        # Declaring an XML object
        tree = ET.parse(file)
        root = tree.getroot()

        for i in root:
            url = i.find('url').text
            domain = i.find('host').text
            status_code = i.find('status').text
            if status_code is not None:
                status_code = int(i.find('status').text)
            else:
                continue

            if 'map?' in url:
                if url.endswith('/') or url.endswith('\\'):
                    url = url[:-1]
                url = url.split('?')[0]

            if url.endswith('.map') and "css.map" not in url and status_code in success_codes:

                if url not in unique_js:
                    unique_js.append(url)
                elif url in unique_js:
                    continue

                response = i.find('response').text

                if response is None:
                    continue

                content_response = base64.b64decode(response)
                content_response = content_response.decode('latin-1')

                try:
                    data = content_response.split('\r\n\r\n', 1)
                    data = data[1]
                except IndexError:
                    data = content_response.split('\n\n', 1)
                    data = data[1]

                if response is None:
                    continue

                sm = get_source_map(url, data)

                if sm is None or not isinstance(sm, dict):
                    print("Invalid sourcemap received, skipping.")
                    continue

                print("Retrieved Sourcemap with version {}, containing {} entries.".format(sm.get('version'),
                                                                                           len(sm.get('sources', []))))

                sources = sm.get('sources')
                sources_content = sm.get('sourcesContent')

                if not sources:
                    print("No sources found.")
                    continue

                if not sources_content:
                    print("No source content found.")
                    continue

                if sm.get('version') != 3:
                    print("[!] Sourcemap is not version 3. This is untested!")

                current_directory = os.getcwd()
                concatenated_path = os.path.join(current_directory, domain)

                os.makedirs(concatenated_path, exist_ok=True)

                for i, source_path in enumerate(sources):
                    source_path = source_path.lstrip('/')
                    source_path = source_path.replace("..", "").replace("/", "\\").replace(" ", "")
                    # If on windows, clean the source path.
                    if os.name == 'nt':
                        source_path = clean_windows(source_path)

                    script_path = concatenated_path + source_path

                    if i < len(sources_content):
                        script_data = sources_content[i]
                        write_file(script_path, script_data)

                    else:
                        print("Missing source content for source path:", source_path)

                print(
                    f"{GREEN}[SUCCESS]{RESET} The JavaScript source code extracted to {BLUE}{concatenated_path}{RESET} directory.{RESET}")

                output_file = f"{concatenated_path}/node_modules.txt"
                import_re = re.compile(r"import.*(/|\.\/|\.\.\/)")
                file_re = re.compile(r"from\s*['\"]([^;'\"]+)['\"]")
                packages = set()

                # Traverse directory
                for dirpath, dirs, files in os.walk(concatenated_path):
                    for filename in files:
                        fname = os.path.join(dirpath, filename)
                        with open(fname, encoding="latin-1") as myfile:
                            lines = myfile.readlines()
                            for line in lines:
                                if "import" in line:
                                    # Check that import statement does not contain /, ./, or ../
                                    if not import_re.search(line):
                                        match = file_re.search(line)
                                        if match:
                                            package = match.group(1).strip().replace(" ", "").replace(";", "")
                                            packages.add(package)

                # Write unique, cleaned-up packages to output_file
                with open(output_file, 'w', encoding="latin-1") as f:
                    for package in packages:
                        if package != "" and package != "." and package != "," and package != ";":
                            f.write(package + "\n")

                # Read the file line by line
                with open(output_file, 'r', encoding="latin-1") as f:
                    for package in f.readlines():
                        package = package.strip()
                        print(f"{YELLOW}[INFO]{RESET} Checking package: {BLUE}'{package}{RESET}'")

                        # Use npm view to check if the package exists
                        try:
                            subprocess.check_output(f"npm view {package}", shell=True)
                            print(f"{YELLOW}[INFO]{RESET} Package {BLUE}{package}{RESET} exists.")
                        except subprocess.CalledProcessError:
                            print(f"{GREEN}[SUCCESS]{GREEN} Package {BLUE}{package}{RESET} does not exist!")
                sys.exit(1)

        if len(unique_js) == 0:
            print(f"\n{RED}[WARN]{RESET} Nothing found, double check your Burp Suite filters for JS extension.")

    elif jsCheck:

        # Declaring an XML object
        tree = ET.parse(file)
        root = tree.getroot()
        packages = set()
        uniqueURLs = []

        # Looping through each request/response
        for i in root:

            # Searching for responses only
            response = i.find('response').text
            if response is None:
                continue

            url = i.find('url').text
            if url.endswith('.js') or 'js?' in url or url.endswith(
                    '.map') or 'map?' in url and ".min." not in url and ".css" not in url:
                print(f'{YELLOW}[INFO]{RESET} Checking: {url}')
                if 'js?' in url or 'map?' in url:
                    if url.endswith('/') or url.endswith('\\'):
                        url = url[:-1]
                    url = url.split('?')[0]
            else:
                continue

            if url in uniqueURLs:
                continue
            else:
                uniqueURLs.append(url)

            # Decoding the response
            response = base64.b64decode(response)
            response = response.decode('latin-1')
            output_file = f"node_modules.txt"

            import_re = re.compile(r"import.*(/|\.\/|\.\.\/)")
            file_re = re.compile(r"from\s*['\"]([^;'\"]+)['\"]")

            if "import { " in response:

                # Check that import statement does not contain /, ./, or ../
                if not import_re.search(response):

                    match = file_re.search(response)

                    if match:
                        package = match.group(1).strip().replace(" ", "").replace(";", "")

                        packages.add(package)

        if len(packages) > 0:
            # Write unique, cleaned-up packages to output_file
            with open(output_file, 'w', encoding="latin-1") as f:
                for package in packages:
                    if package != "" and package != "." and package != "," and package != ";":
                        f.write(package + "\n")

            # Read the file line by line
            with open(output_file, 'r', encoding="latin-1") as f:
                for package in f.readlines():
                    package = package.strip()
                    print(f"Checking package: '{package}'")

                    # Use npm view to check if the package exists
                    try:
                        subprocess.check_output(f"npm view {package}", shell=True)
                        print(f"{RED}[WARN]{RESET} Package {BLUE}{package}{RESET} exists.")
                    except subprocess.CalledProcessError:
                        print(f"{GREEN}[SUCCESS]{RESET} Package {BLUE}{package}{RESET} does not exist!")
                sys.exit(1)
        else:
            print(f"\n{RED}[WARN]{RESET} Nothing found, double check your Burp Suite filters for JS extension.")


if __name__ == '__main__':

    try:
        args = parse_args()
    except SystemExit:
        update_notifier()
        sys.exit()

    if args.update:
        check_for_updates()
        sys.exit(1)

    if args.version:
        version = get_current_version()
        print("\n" + version)
        sys.exit(0)

    start = time.time()

    if not args.wordlist and not args.postoexcel and not args.dependency and "-h" not in sys.argv and "--help" not in sys.argv:
        # Create a multiprocessing manager
        manager = multiprocessing.Manager()

        # Create a shared list using the manager
        final_xlsx = manager.list()

        # Get the number of CPU cores
        num_processes = args.threads

        # Create a pool of processes
        pool = multiprocessing.Pool(processes=num_processes)

    # Some regex for finding intersting stuff
    regex_secrets = {
        'google_api': r'AIza[0-9A-Za-z\-_]{35}',
        'firebase': r'AAAA[A-Za-z0-9_\-]{7}:[A-Za-z0-9_\-]{140}',
        'google_oauth': r'ya29\.[0-9A-Za-z\\\-_]+',
        'amazon_aws_access_key_id': r'A[SK]IA[0-9A-Z]{16}',
        'amazon_mws_auth_toke': r'amzn\\.mws\\.[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}',
        'amazon_aws_url': r's3\.amazonaws\.com[/]+|[a-zA-Z0-9_\-]*\.s3\.amazonaws\.com',
        'amazon_aws_url2': r"([a-zA-Z0-9\-\.\_]+\.s3\.amazonaws\.com)",
        'amazon_aws_url3': r"|s3://[a-zA-Z0-9\-\.\_]+",
        'amazon_aws_url4': r"|s3-[a-zA-Z0-9\-\.\_\/]+",
        'amazon_aws_url5': r"|s3\.amazonaws\.com/[a-zA-Z0-9\-\.\_]+",
        # 'aws_s3_bucket1': r'\b(?:[a-z0-9.-]{3,255}|s3\.console\.aws\.amazon\.com/s3/buckets/[a-zA-Z0-9\-._]+)\b',
        'facebook_access_token': r'EAACEdEose0cBA[0-9A-Za-z]+',
        'authorization_basic': r'basic [a-zA-Z0-9=:_\+\/\-]{5,100}',
        'authorization_bearer': r'bearer [a-zA-Z0-9_\\\-\.=:_\+\/]{5,100}',
        'authorization_api': r'api\[key\|_key\|\s+\]+[a-zA-Z0-9_\\\-]{5,100}',
        'paypal_braintree_access_token': r'access_token\$production\$[0-9a-z]{16}\$[0-9a-f]{32}',
        'square_oauth_secret': r'sq0csp\-[ 0-9A-Za-z\\\-_]{43}|sq0[a-z]{3}\-[0-9A-Za-z\\\-_]{22,43}',
        'square_access_token': r'sqOatp\-[0-9A-Za-z\\\-_]{22}|EAAA[a-zA-Z0-9]{60}',
        'stripe_standard_api': r'sk_live_[0-9a-zA-Z]{24}',
        'stripe_restricted_api': r'rk_live_[0-9a-zA-Z]{24}',
        'github_access_token': r'[a-zA-Z0-9_\\\-]*:[a-zA-Z0-9_\\\-]+@github\.com*',
        'rsa_private_key': r'-----BEGIN RSA PRIVATE KEY-----',
        'ssh_dsa_private_key': r'-----BEGIN DSA PRIVATE KEY-----',
        'ssh_dc_private_key': r'-----BEGIN EC PRIVATE KEY-----',
        'pgp_private_block': r'-----BEGIN PGP PRIVATE KEY BLOCK-----',
        'json_web_token': r'ey[A-Za-z0-9_\\\-=\+\/]{5,100}\.[A-Za-z0-9_\\\-=\+\/]{5,100}\.?[A-Za-z0-9_\\\-=\+\/]*$',
        'slack_token': r'"api_token":"(xox[a-zA-Z]\-[a-zA-Z0-9\-]+)"',
        'SSH_privKey': r"([-]+BEGIN [^\s]+ PRIVATE KEY[-]+[\s]*[^-]*[-]+END [^\s]+ PRIVATE KEY[-]+)",
        'email_address': r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b',
        'aws_access_key': r'AKIA[0-9A-Z]{16}',
        'ssh_private_key': r'-----BEGIN (?:RSA|DSA|EC|OPENSSH) PRIVATE KEY-----[a-zA-Z0-9+\/=\s]+-----END (?:RSA|DSA|EC|OPENSSH) PRIVATE KEY-----',
        'md5_hash': r'\b[a-f0-9]{32}\b',
        'sha1_hash': r'\b[a-f0-9]{40}\b',
        'sha256_hash': r'\b[a-f0-9]{64}\b',
        'ssn': r'\b[0-9]{3}\-[0-9]{2}\-[0-9]{4}\b',
        'ccn': r'\b(?:\d[ -]*?){13,16}\b',
        # 'api_token': r'\b[A-Za-z0-9_]{32}\b',
        'aws_s3_bucket': r'\b[a-z0-9.-]{3,255}\b',
        'aws_secret_access_key': r'[A-Za-z0-9/+=]{40}',
        'slack_webhook_url': r'https://hooks\.slack\.com/services/[A-Z0-9/]+',
        'ssn_with_dashes': r'\b[0-9]{3}\-[0-9]{2}\-[0-9]{4}\b',
        'github_access_token_v2': r'[a-f0-9]{40}',
        'mongodb_connection_string': r'mongodb\+srv://[^:\s]+:[^@\s]+@[^/\s]+/\S+',
        'phone_number': r'\b(?:\+\d{1,3}[-. ]?)?\(?\d{3}\)?[-. ]?\d{3}[-. ]?\d{4}\b',
        'cc_expiration': r'\b\d{2}\/\d{2}\b',
        'bitcoin_address': r'\b[13][a-km-zA-HJ-NP-Z1-9]{25,34}\b',
    }

    # Regex for finding Paths and APIs Endpoints *
    # You can uncomment if you want to find more complex hardcoded path and APIs
    api_extractor = {
        'OpenAI Generator1': r'\/v1\/(.+)',
        'OpenAI Generator2': r'\/v2\/(.+)',
        # 'PATH Finder-v1': r'\/[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        # 'PATH Finder-v2': r'[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        'PATH Finder-v3': r'\/api[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        'PATH Finder-v4': r'api[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        # 'PATH Finder-v5': r'\/[a-zA-Z]+(?:\/[a-zA-Z]+)+\/(?=[a-zA-Z]+\/)?[a-zA-Z]+(?:\/[a-zA-Z]+)*',
        # 'PATH Finder-v6': r'\/[a-zA-Z0-9]+(?:\/[a-zA-Z-0-9]+)*(?:\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
        # 'PATH Finder-v7': r'[a-zA-Z0-9]+(?:\/[a-zA-Z-0-9]+)*(?:\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
        # 'PATH Finder Backslash-1': r'\/[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        # 'PATH Finder Backslash-v2': r'[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        'PATH Finder Backslash-v3': r'\/api[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        'PATH Finder Backslash-v4': r'api[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
        # 'PATH Finder Backslash-v5': r'\\\/[a-zA-Z]+(?:\\\/[a-zA-Z]+)+\\\/(?=[a-zA-Z]+\\\/)?[a-zA-Z]+(?:\\\/[a-zA-Z]+)*',
        # 'PATH Finder Backslash-v6': r'\\\/[a-zA-Z0-9]+(?:\\\/[a-zA-Z-0-9]+)*(?:\\\\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
        # 'PATH Finder Backslash-v7': r'[a-zA-Z]+(?:\\\/[a-zA-Z]+)+\\\/(?=[a-zA-Z]+\\\/)?[a-zA-Z]+(?:\\\/[a-zA-Z]+)*',
        # 'PATH Finder Backslash-v8': r'[a-zA-Z0-9]+(?:\\\/[a-zA-Z-0-9]+)*(?:\\\\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
    }

    # Regex for finding URLs
    uri_finder = {
        'URL Finder': '(http|ftp|https):\/\/([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^=%&:\/~+#-]*[\w@?^=%&\/~+#-])',
        'URL Finder Backslash': '(http|ftp|https):\\\/\\\/([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^=%&:\\\/~+#-]*[\w@?^=%&\\\/~+#-])'
    }

    warnings.filterwarnings("ignore")

    files = []
    matched_patterns = set()
    path_js = set()
    task_args = []

    # Creating a new Workbook object
    wb = Workbook()
    wb_json = Workbook()

    if args.all and args.secrets and args.api and args.urls and args.path and args.json and args.js and args.wordlist and args.map and args.paramspider:
        print(
            f'\n{RED}[WARN]{RESET} If --all is set, remove other arguments(api/secrets/urls/json/postman/js/domain).{RESET}')
        exit(1)

    if args.postoexcel:

        answerOne = input(
            f"{YELLOW}[Q]{RESET} Would you like to supply a directory path containing all Postman files or a single file path? {BLUE}F/D (F for file D for directory){RESET}: ")

        if answerOne.upper().strip() == "D":
            json_directory = input(f"{YELLOW}[INFO]{RESET} Enter the directory path containing all Postman files: ")

            answerTwo = input(
                f"{YELLOW}[Q]{RESET} Do you wish to print the final count of each HTTP method? {GREEN}Y{RESET}/{RED}N{RESET}: ")
            if answerTwo.upper().strip() == "Y":

                postmanDirectory(json_directory.strip(), True)

                answerTwoLoop = input(
                    f"{YELLOW}[Q]{RESET} Do you wish to export all endpoints to Excel sheets as well? {GREEN}Y{RESET}/{RED}N{RESET}: ")
                if answerTwoLoop.upper().strip() == "Y":
                    postmanDirectory(json_directory.strip())

            else:
                answerThree = input(
                    f"{YELLOW}[Q]{RESET} Do you wish to export all endpoints to Excel sheets? {GREEN}Y{RESET}/{RED}N{RESET}: ")
                if answerThree.upper().strip() == "Y":
                    postmanDirectory(json_directory.strip())

                else:
                    exit(1)

        elif answerOne.upper().strip() == "F":
            file_path = input(f"{YELLOW}[Q]{RESET} Enter the file path of the Postman file: {RESET}")

            answerTwo = input(
                f"{YELLOW}[Q]{RESET} Do you wish to print the final count of each HTTP method? {GREEN}Y{RESET}/{RED}N{RESET}: ")
            if answerTwo.upper().strip() == "Y":
                postmanFile(file_path.strip(), True)

                answerTwoLoop = input(
                    f"{YELLOW}[Q]{RESET} Do you wish to export all endpoints to Excel sheets as well? {GREEN}Y{RESET}/{RED}N{RESET}: ")
                if answerTwoLoop.upper().strip() == "Y":
                    postmanFile(file_path.strip())

            elif answerTwo.upper().strip() == "N":
                answerThree = input(
                    f"{YELLOW}[Q]{RESET} Do you wish to export all endpoints to Excel sheets? {GREEN}Y{RESET}/{RED}N{RESET}: ")
                if answerThree.upper().strip() == "Y":
                    postmanFile(file_path.strip())
                else:
                    exit(1)

    if args.directory and args.file:
        print(f'\n{RED}[WARN]{RESET} Choose either --file or --directory not both!')
        exit(1)

    elif args.file:
        filename = args.file

    elif args.directory:
        directory = os.fsencode(args.directory)
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            files.append(os.path.join(directory.decode(), filename))
            if "." in filename or filename.endswith(".py") or filename.endswith(".txt"):
                continue

    if args.all and args.secrets and args.api and args.urls and args.path and args.json and args.js and args.wordlist and args.map and args.paramspider:
        print(
            f'\n{RED}[WARN]{RESET} If --all is set, remove other arguments(api/secrets/urls/json/postman/js/domain).{RESET}')
        exit(1)

    if args.all:

        if args.directory:

            for filename in files:

                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                sheet_secrets, wb_secrets = create_worksheet(host, "Secrets")

                sheet_api_finder, wb_api_finder = create_worksheet(host, "Path_and_Endpoints")

                sheet_url_finder, wb_url_finder = create_worksheet(host, "URLs")

                sheet_sub_domains, wb_sub_domains = create_worksheet(host, "Sub-Domains")

                sheet_js, wb_js = create_worksheet_js("JS-Files")

                if not args.verbose:
                    print(
                        f'\n{YELLOW}[INFO]{RESET} Executing APIs Collector, Postman, JSON files, Secrets, URLs, APIs in JS files and Sub-domains features for {BLUE}{host}{RESET}.')
                    print(f'{YELLOW}[INFO]{RESET} Add --verbose to see the output printed to the screen with colors.')
                    print(f'\n{YELLOW}[INFO]{RESET} This might take a while, be patient I tell you!')

                print(f'\n{YELLOW}[INFO]{RESET} Executing  feature for {BLUE}{host}{RESET}.')
                api_collector(filename)
                print(f'\n{YELLOW}[INFO]{RESET} Executing Postman feature for {BLUE}{host}{RESET}.')
                postMan(filename)
                print(f'\n{YELLOW}[INFO]{RESET} Executing JSON feature for {BLUE}{host}{RESET}.')
                json_file(filename, wb_json)
                print(f'\n{YELLOW}[INFO]{RESET} Executing JS URLs feature for {BLUE}{host}{RESET}.')
                js_file(filename, wb_js, sheet_js)
                print(f'\n{YELLOW}[INFO]{RESET} Executing URLs feature for {BLUE}{host}{RESET}.')
                task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))
                print(f'\n{YELLOW}[INFO]{RESET} Executing Secrets feature for {BLUE}{host}{RESET}.')
                task_args.append((filename, "Secrets", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))
                print(f'\n{YELLOW}[INFO]{RESET} Executing APIs & Paths feature for {BLUE}{host}{RESET}.')
                task_args.append((filename, "Path_and_Endpoints", sheet_url_finder, wb_url_finder, uri_finder,
                                  regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
                print(f'\n{YELLOW}[INFO]{RESET} Executing Sub-Domains feature for {BLUE}{host}{RESET}.')
                task_args.append((filename, "Sub-Domains", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))
                print(f'\n{YELLOW}[INFO]{RESET} Creating a wordlist tailored to {BLUE}{host}{RESET}.')
                wordlist_creator(filename, host)
                print(f'\n{YELLOW}[INFO]{RESET} Converting map files to their original source code for.')
                mapConverter(filename)
                print(f'\n{YELLOW}[INFO]{RESET} Executing paramspider.')
                paramspider(filename)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.verbose:
                print(
                    f'\n{YELLOW}[INFO]{RESET} Executing APIs Collector, Postman, JSON files, Secrets, URLs, APIs in JS files and Sub-domains features for {BLUE}{host}{RESET}.')
                print(f'{YELLOW}[INFO]{RESET} Add --verbose to see the output printed to the screen with colors.')
                print(f'\n{YELLOW}[INFO]{RESET} This might take a while, be patient I tell you!')

            sheet_url_finder, wb_url_finder = create_worksheet_main("URLs")

            sheet_secrets, wb_secrets = create_worksheet_main("Secrets")

            sheet_api_finder, wb_api_finder = create_worksheet_main("API Endpoints")

            sheet_sub_domains, wb_sub_domains = create_worksheet_main("Sub Domains")

            sheet_js, wb_js = create_worksheet_js("JS-Files")

            print(f'\n{YELLOW}[INFO]{RESET} Executing APIs Collector feature for {BLUE}{host}{RESET}.')
            api_collector(filename)
            print(f'\n{YELLOW}[INFO]{RESET} Executing Postman feature for {BLUE}{host}{RESET}.')
            postMan(filename)
            print(f'\n{YELLOW}[INFO]{RESET} Executing JSON feature for {BLUE}{host}{RESET}.')
            json_file(filename, wb_json)
            print(f'\n{YELLOW}[INFO]{RESET} Executing JS URLs feature for {BLUE}{host}{RESET}.')
            js_file(filename, wb_js, sheet_js)
            print(f'\n{YELLOW}[INFO]{RESET} Executing URLs feature for {BLUE}{host}{RESET}.')
            task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                              api_extractor, args, matched_patterns, final_xlsx))
            print(f'\n{YELLOW}[INFO]{RESET} Executing Secrets feature for {BLUE}{host}{RESET}.')
            task_args.append((filename, "Secrets", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                              api_extractor, args, matched_patterns, final_xlsx))
            print(f'\n{YELLOW}[INFO]{RESET} Executing APIs & Paths feature for {BLUE}{host}{RESET}.')
            task_args.append((
                             filename, "Path_and_Endpoints", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                             api_extractor, args, matched_patterns, final_xlsx))
            print(f'\n{YELLOW}[INFO]{RESET} Executing Sub-Domains feature for {BLUE}{host}{RESET}.')
            task_args.append((filename, "Sub-Domains", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                              api_extractor, args, matched_patterns, final_xlsx))
            print(f'\n{YELLOW}[INFO]{RESET} Creating a wordlist tailored to {BLUE}{host}{RESET}.')
            wordlist_creator(filename, host)
            print(f'\n{YELLOW}[INFO]{RESET} Executing paramspider.')
            paramspider(filename)

    if not args.all:
        sheet_url_finder, wb_url_finder = create_worksheet_main("URLs")

        sheet_secrets, wb_secrets = create_worksheet_main("Secrets")

        sheet_api_finder, wb_api_finder = create_worksheet_main("API Endpoints")

        sheet_sub_domains, wb_sub_domains = create_worksheet_main("Sub Domains")

        sheet_js, wb_js = create_worksheet_js("JS-Files")

    if args.api and not args.all:

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing APIs collector feature for {BLUE}{host}{RESET}.')

                api_collector(filename)
                    
                update_notifier()

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing APIs collector feature for {BLUE}{host}{RESET}.')

            api_collector(filename)
    
            update_notifier()

    if args.paramspider and not args.all:

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing paramspider feature for {BLUE}{host}{RESET}.')

                paramspider(filename)

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing paramspider feature for {BLUE}{host}{RESET}.')

            paramspider(filename)

    if args.postman and not args.all:
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing Postman feature for {BLUE}{host}{RESET}.')

                postMan(filename)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing Postman feature for {BLUE}{host}{RESET}.')

            postMan(filename)

    if args.map and not args.all:

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing Map converter feature for {BLUE}{host}{RESET}.')

                mapConverter(filename)

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing Map converter feature for {BLUE}{host}{RESET}.')

            mapConverter(filename)

    if args.dependency and not args.all:

        jsCheck = input(
            f"{YELLOW}[Q]{RESET} Do you wish to check the Javascript files first? {GREEN}Y{RESET}/{RED}N{RESET}: ")
        if jsCheck.upper() == 'N':
            jsCheck = None
        elif jsCheck.upper() == 'Y':
            jsCheck = True
        else:
            print(f"{RED}[WARN]{RESET} You must choose either Y or N!")
            sys.exit(1)

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing Dependency check feature for {BLUE}{host}{RESET}.')

                dependencyCheck(filename, jsCheck)

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing Dependency check feature for {BLUE}{host}{RESET}.')

            dependencyCheck(filename, jsCheck)

    if args.js and not args.all:
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing JS URLs feature for {BLUE}{host}{RESET}.')

                js_file(filename, wb_js, sheet_js)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing JS URLs feature for {BLUE}{host}{RESET}.')

            js_file(filename, wb_js, sheet_js)

    if args.urls and not args.all:

        counter = 0
        if args.directory:
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing URLs feature for {BLUE}{host}{RESET}.')

                sheet_url_finder, wb_url_finder = create_worksheet_main(f"URLs_{counter}")

                task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))


        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing URLs feature for {BLUE}{host}{RESET}.')

            task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, regex_secrets,
                              api_extractor, args, matched_patterns, final_xlsx))

    if args.domain and not args.all:

        if args.directory:
            counter = 0
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing Sub-Domains feature for {BLUE}{host}{RESET}.')

                sheet_sub_domains, wb_sub_domains = create_worksheet_main(f"Sub Domains_{counter}")
                # Map the function and arguments to the pool
                task_args.append((filename, "Sub-Domains", sheet_sub_domains, wb_sub_domains, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))


        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing Sub-Domains feature for {BLUE}{host}{RESET}.')

                # Map the function and arguments to the pool
            task_args.append((filename, "Sub-Domains", sheet_sub_domains, wb_sub_domains, uri_finder, regex_secrets,
                              api_extractor, args, matched_patterns, final_xlsx))

    if args.json and not args.all:
        # Creating a new Workbook object
        wb_json = Workbook()

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing JSON feature for {BLUE}{host}{RESET}.')

                json_file(filename, wb_json)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing JSON feature for {BLUE}{host}{RESET}.')

            json_file(filename, wb_json)

    if args.path and not args.all:

        if args.directory:

            counter = 0
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} This might take a while, be patient I tell you!')
                    print(
                        f'\n{YELLOW}[INFO]{RESET} Executing APIs and PATHs with REGEX feature for {BLUE}{host}{RESET}.')

                sheet_api_finder, wb_api_finder = create_worksheet_main(f"API Endpoints_{counter}")

                # Map the function and arguments to the pool
                task_args.append((filename, "Path_and_Endpoints", sheet_api_finder, wb_api_finder, uri_finder,
                                  regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing APIs and PATHs with REGEX feature for {BLUE}{host}{RESET}.')
                print(f'\n{YELLOW}[INFO]{RESET} This might take a while, be patient I tell you!')

            # Map the function and arguments to the pool
            task_args.append((
                             filename, "Path_and_Endpoints", sheet_api_finder, wb_api_finder, uri_finder, regex_secrets,
                             api_extractor, args, matched_patterns, final_xlsx))

    if args.secrets and not args.all:

        counter = 0
        argumnets = [args.urls, args.path, args.domain]
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:
                    print(f'\n{YELLOW}[INFO]{RESET} Executing Secrets feature with REGEX for {BLUE}{host}{RESET}.')

                sheet_secrets, wb_secrets = create_worksheet_main(f"Secrets_{counter}")
                # Map the function and arguments to the pool
                task_args.append((filename, "Secrets", sheet_secrets, wb_secrets, uri_finder, regex_secrets,
                                  api_extractor, args, matched_patterns, final_xlsx))

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:
                print(f'\n{YELLOW}[INFO]{RESET} Executing Secrets feature with REGEX for {BLUE}{host}{RESET}.')

            task_args.append((filename, "Secrets", sheet_secrets, wb_secrets, uri_finder, regex_secrets, api_extractor,
                              args, matched_patterns, final_xlsx))

    if args.wordlist and not args.all:

        # Create an XML Object
        tree = ET.parse(filename)
        main_root = tree.getroot()

        for i in main_root:
            response = i.find('response').text
            if response is None:
                continue
            content = base64.b64decode(response)
            content = content.decode('latin-1')
            host = i.find('host').text
            break

        if not args.all and not args.verbose:
            print(f'\n{YELLOW}[INFO]{RESET} Creating a wordlist tailored to {BLUE}{host}{RESET}.')

        wordlist_creator(filename, host)

    if not args.wordlist and not args.postoexcel and not args.dependency and not args.paramspider:
        # Create a pool of processes using a context manager
        with multiprocessing.Pool(processes=num_processes) as pool:
            # Start the processes
            results = pool.starmap_async(main, task_args)

            # Wait for the results 
            while not results.ready():
                time.sleep(0)

        # Terminate all processes
        pool.terminate()
        pool.join()

        check_dup = []

        for excel in final_xlsx:
            if excel not in check_dup:
                check_dup.append(excel)
                adjust_column_widths_disk(excel)

            if excel in check_dup:
                continue

    end = time.time()
    hours, rem = divmod(end - start, 3600)
    minutes, seconds = divmod(rem, 60)
    print("\n--- Running time: {:0>2}:{:0>2}:{:05.2f}".format(int(hours), int(minutes), seconds), "---")
